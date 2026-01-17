# app.py – Professional Streamlit Invoice Model Merger

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Page config ----------
st.set_page_config(
    page_title="Invoice Processor",
    layout="wide",
    page_icon="📦",
    initial_sidebar_state="expanded",
)

# ---------- Custom CSS ----------
st.markdown(
    """
    <style>
    .main-title {
        font-size: 36px;
        font-weight: 800;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 10px;
    }
    .subtitle {
        color: #cccccc;
        font-size: 16px;
        margin-bottom: 20px;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 12px;
        margin: 10px 0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<p class="main-title">📦 Invoice Model Merger</p>', unsafe_allow_html=True)
st.markdown(
    '<p class="subtitle">Koi bhi invoice Excel upload karo. App MODEL ke hisaab se '
    'duplicate rows merge karega, saare columns safe rahenge, aur professional Excel '
    'output dega.</p>',
    unsafe_allow_html=True,
)
st.markdown("---")


# ---------- Helper: detect column ----------
def detect_column(df_cols, keywords):
    """Try to find a column whose name contains any of the given keywords."""
    df_cols_upper = [str(c).upper() for c in df_cols]
    for key in keywords:
        key_up = key.upper()
        for col, col_up in zip(df_cols, df_cols_upper):
            if key_up in col_up:
                return col
    return None


# ---------- Main app ----------
uploaded_file = st.file_uploader(
    "Upload your invoice Excel (.xlsx)", type=["xlsx"]
)

if uploaded_file is not None:
    # 1) Read Excel
    try:
        df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"❌ File error: {e}")
        st.stop()

    st.markdown(
        f'<div class="success-box">✅ File loaded: <b>{uploaded_file.name}</b> '
        f'| Rows: {len(df)}</div>',
        unsafe_allow_html=True,
    )

    with st.expander("🔍 Preview: Original data (first 10 rows)", expanded=False):
        st.dataframe(df.head(10), use_container_width=True)

    # 2) Column names
    original_cols = df.columns.copy()

    # 3) Auto-detect important columns (use original names)
    model_col = detect_column(original_cols, ["MODEL", "STYLE", "ITEM", "CODE"])
    qty_col = detect_column(original_cols, ["QTY", "QUANTITY", "PCS"])
    price_col = detect_column(original_cols, ["PRICE", "U.PRICE", "UNIT"])
    amount_col = detect_column(original_cols, ["AMOUNT", "TOTAL", "VALUE"])

    if not model_col:
        st.error("❌ MODEL / STYLE column nahi mil raha! Column naam check karo.")
        st.stop()

    # ---------- Sidebar ----------
    st.sidebar.header("⚙️ Settings")
    st.sidebar.write("**Detected columns:**")
    st.sidebar.info(
        f"🔹 Model: **{model_col}**\n\n"
        f"🔹 Qty: **{qty_col if qty_col else 'Not found'}**\n\n"
        f"🔹 Price: **{price_col if price_col else 'Not found'}**\n\n"
        f"🔹 Amount: **{amount_col if amount_col else 'Not found'}**"
    )

    st.sidebar.markdown("---")
    st.sidebar.write("**Result filters:**")
    min_qty = st.sidebar.number_input("Min Total QTY", value=0, min_value=0, step=1)
    search_model = st.sidebar.text_input("Filter model (contains)", value="")
    sort_by_options = ["MODEL"]
    if qty_col:
        sort_by_options.append("Total_QTY")
    if amount_col:
        sort_by_options.append("Total_Amount")
    sort_by = st.sidebar.selectbox("Sort by", options=sort_by_options, index=0)

    # 4) Clean data
    working_df = df.copy()

    # Clean MODEL column
    working_df[model_col] = working_df[model_col].astype(str).str.strip()
    working_df[model_col] = working_df[model_col].replace(
        ["", "nan", "NaN", "NONE", "None"], pd.NA
    )
    working_df[model_col] = working_df[model_col].fillna(method="ffill")

    # Convert numeric columns
    if qty_col:
        working_df[qty_col] = pd.to_numeric(
            working_df[qty_col], errors="coerce"
        ).fillna(0)
    if price_col:
        working_df[price_col] = pd.to_numeric(
            working_df[price_col], errors="coerce"
        ).fillna(0)
    if amount_col:
        working_df[amount_col] = pd.to_numeric(
            working_df[amount_col], errors="coerce"
        ).fillna(0)

    # 5) Groupby MODEL – sum QTY, AMOUNT; keep first of others
    agg_dict = {}
    if qty_col:
        agg_dict[qty_col] = "sum"
    if price_col:
        agg_dict[price_col] = "first"
    if amount_col:
        agg_dict[amount_col] = "sum"

    for col in working_df.columns:
        if col != model_col and col not in agg_dict:
            agg_dict[col] = "first"

    grouped = working_df.groupby(model_col, as_index=False).agg(agg_dict)  # [web:47]

    # 6) Helper columns for display
    display_df = grouped.copy()
    if qty_col:
        display_df["Total_QTY"] = display_df[qty_col]
    if amount_col:
        display_df["Total_Amount"] = display_df[amount_col]

    # 7) Filters
    if min_qty > 0 and qty_col:
        display_df = display_df[display_df["Total_QTY"] >= min_qty]

    if search_model:
        display_df = display_df[
            display_df[model_col]
            .astype(str)
            .str.contains(search_model, case=False, na=False)
        ]

    # 8) Sorting
    if sort_by == "Total_QTY" and qty_col:
        sort_col = "Total_QTY"
        ascending = False
    elif sort_by == "Total_Amount" and amount_col:
        sort_col = "Total_Amount"
        ascending = False
    else:
        sort_col = model_col
        ascending = True

    display_df = display_df.sort_values(sort_col, ascending=ascending).reset_index(
        drop=True
    )

    # 9) Metrics + table
    st.markdown("### ✅ Merged Result")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Total Models", len(display_df))
    with c2:
        if qty_col:
            st.metric("Total Quantity", int(display_df["Total_QTY"].sum()))
    with c3:
        if amount_col:
            st.metric(
                "Total Amount",
                f"{display_df['Total_Amount'].sum():,.2f}",
            )

    st.dataframe(display_df, use_container_width=True, height=400)

    # 10) Excel export
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        display_df.to_excel(writer, sheet_name="Merged Data", index=False)

        ws = writer.sheets["Merged Data"]

        # Header style
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border

        # Data rows – alternating colors
        light_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        for row_idx, row in enumerate(
            ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
            ),
            start=2,
        ):
            fill = light_fill if (row_idx - 2) % 2 == 0 else white_fill
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"

        # Auto column widths
        for column in ws.columns:
            max_len = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"

    excel_bytes = output.getvalue()

    st.markdown("### ⬇️ Download")
    st.download_button(
        label="📥 Download Professional Excel",
        data=excel_bytes,
        file_name=f"Invoice_Merged_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("---")
    st.success("✨ Processing complete! Download karke Odoo ya kahin bhi use karo.")

else:
    st.info("📤 Pehle upar se Excel (.xlsx) file upload karo.")
